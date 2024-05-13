# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'C:\Users\White Lightning\Desktop\BuyBackPlus.ui'
#
# Created by: PyQt5 UI code generator 5.13.0
#
# WARNING! All changes made in this file will be lost!

import threading
import lxml
import urllib.request
from urllib.request import Request, urlopen
import json
from bs4 import BeautifulSoup
import requests
from requests import cookies
import time
import winsound
from PyQt5 import QtCore, QtGui, QtWidgets
from selenium.webdriver.support import expected_conditions as ec
from pynput.keyboard import Key, Listener
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from openpyxl import Workbook, load_workbook
import atexit
import pandas as pd

class Ui_BuyBackPlus(object):

    powellloc = "./buylists/powell.xlsx"
    powelldb = load_workbook(powellloc)
    powelldbs = powelldb.active
    booksrunloc = "./buylists/booksrun.xlsx"
    booksrundb = load_workbook(booksrunloc)
    booksrundbs = booksrundb.active
    ecampusloc = "./buylists/ecampus.xlsx"
    ecampusdb = load_workbook(ecampusloc)
    ecampusdbs = ecampusdb.active
    bookmonsterloc = "./buylists/bookmonster.xlsx"
    bookmonsterdb = load_workbook(bookmonsterloc)
    bookmonsterdbs = bookmonsterdb.active
    sbybloc = "./buylists/sbyb.xlsx"
    sbybdb = load_workbook(sbybloc)
    sbybdbs = sbybdb.active
    tbrecloc = "./buylists/tbrec.xlsx"
    tbrecdb = load_workbook(tbrecloc)
    tbrecdbs = tbrecdb.active
    ziffitloc = "./buylists/ziffit.xlsx"
    ziffitdb = load_workbook(ziffitloc)
    ziffitdbs = ziffitdb.active

    location = "./history.xlsx"
    database = load_workbook(location)
    databasesheet = database.active

    def setupUi(self, BuyBackPlus):
        BuyBackPlus.setObjectName("BuyBackPlus")
        BuyBackPlus.setEnabled(True)
        BuyBackPlus.resize(690, 420)
        BuyBackPlus.setSizeIncrement(QtCore.QSize(0, 0))
        self.centralwidget = QtWidgets.QWidget(BuyBackPlus)
        self.centralwidget.setObjectName("centralwidget")
        self.titlelabel = QtWidgets.QLabel(self.centralwidget)
        self.titlelabel.setGeometry(QtCore.QRect(350, 10, 321, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.titlelabel.setFont(font)
        self.titlelabel.setAlignment(QtCore.Qt.AlignCenter)
        self.titlelabel.setObjectName("titlelabel")
        self.isbnlabel = QtWidgets.QLabel(self.centralwidget)
        self.isbnlabel.setGeometry(QtCore.QRect(350, 50, 321, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.isbnlabel.setFont(font)
        self.isbnlabel.setAlignment(QtCore.Qt.AlignCenter)
        self.isbnlabel.setObjectName("isbnlabel")
        self.topbuyback = QtWidgets.QLabel(self.centralwidget)
        self.topbuyback.setGeometry(QtCore.QRect(20, 20, 301, 51))
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.topbuyback.setFont(font)
        self.topbuyback.setAlignment(QtCore.Qt.AlignCenter)
        self.topbuyback.setObjectName("topbuyback")
        self.firstlabel = QtWidgets.QLabel(self.centralwidget)
        self.firstlabel.setGeometry(QtCore.QRect(20, 70, 301, 51))
        font = QtGui.QFont()
        font.setFamily("Nirmala UI")
        font.setPointSize(16)
        font.setBold(False)
        font.setWeight(50)
        self.firstlabel.setFont(font)
        self.firstlabel.setAutoFillBackground(False)
        self.firstlabel.setStyleSheet("background-color: rgb(161, 255, 19);")
        self.firstlabel.setAlignment(QtCore.Qt.AlignCenter)
        self.firstlabel.setObjectName("firstlabel")
        self.secondbuyback = QtWidgets.QLabel(self.centralwidget)
        self.secondbuyback.setGeometry(QtCore.QRect(20, 120, 301, 51))
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.secondbuyback.setFont(font)
        self.secondbuyback.setAlignment(QtCore.Qt.AlignCenter)
        self.secondbuyback.setObjectName("secondbuyback")
        self.secondlabel = QtWidgets.QLabel(self.centralwidget)
        self.secondlabel.setGeometry(QtCore.QRect(20, 170, 301, 51))
        font = QtGui.QFont()
        font.setFamily("Nirmala UI")
        font.setPointSize(16)
        font.setBold(False)
        font.setWeight(50)
        self.secondlabel.setFont(font)
        self.secondlabel.setStyleSheet("background-color: rgb(247, 255, 76);")
        self.secondlabel.setAlignment(QtCore.Qt.AlignCenter)
        self.secondlabel.setObjectName("secondlabel")
        self.thirdbuyback = QtWidgets.QLabel(self.centralwidget)
        self.thirdbuyback.setGeometry(QtCore.QRect(20, 220, 301, 51))
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.thirdbuyback.setFont(font)
        self.thirdbuyback.setAlignment(QtCore.Qt.AlignCenter)
        self.thirdbuyback.setObjectName("thirdbuyback")
        self.thirdlabel = QtWidgets.QLabel(self.centralwidget)
        self.thirdlabel.setGeometry(QtCore.QRect(20, 270, 301, 51))
        font = QtGui.QFont()
        font.setFamily("Nirmala UI")
        font.setPointSize(16)
        font.setBold(False)
        font.setWeight(50)
        self.thirdlabel.setFont(font)
        self.thirdlabel.setStyleSheet("background-color: rgb(255, 159, 111);")
        self.thirdlabel.setAlignment(QtCore.Qt.AlignCenter)
        self.thirdlabel.setObjectName("thirdlabel")
        self.otherlabel = QtWidgets.QLabel(self.centralwidget)
        self.otherlabel.setGeometry(QtCore.QRect(350, 170, 321, 151))
        font = QtGui.QFont()
        font.setFamily("Nirmala UI")
        font.setPointSize(16)
        font.setBold(False)
        font.setWeight(50)
        self.otherlabel.setFont(font)
        self.otherlabel.setAutoFillBackground(False)
        self.otherlabel.setStyleSheet("background-color: rgb(235, 90, 61);")
        self.otherlabel.setAlignment(QtCore.Qt.AlignCenter)
        self.otherlabel.setObjectName("otherlabel")
        self.otherbuyback = QtWidgets.QLabel(self.centralwidget)
        self.otherbuyback.setGeometry(QtCore.QRect(350, 120, 321, 51))
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.otherbuyback.setFont(font)
        self.otherbuyback.setAlignment(QtCore.Qt.AlignCenter)
        self.otherbuyback.setObjectName("otherbuyback")
        self.offerlabel = QtWidgets.QLabel(self.centralwidget)
        self.offerlabel.setGeometry(QtCore.QRect(350, 90, 321, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.offerlabel.setFont(font)
        self.offerlabel.setAlignment(QtCore.Qt.AlignCenter)
        self.offerlabel.setObjectName("offerlabel")
        self.statuslabel = QtWidgets.QLabel(self.centralwidget)
        self.statuslabel.setGeometry(QtCore.QRect(20, 340, 651, 51))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.statuslabel.setFont(font)
        self.statuslabel.setAlignment(QtCore.Qt.AlignCenter)
        self.statuslabel.setObjectName("statuslabel")
        BuyBackPlus.setCentralWidget(self.centralwidget)
        self.scannedlabel = QtWidgets.QLabel(self.centralwidget)
        self.scannedlabel.setGeometry(QtCore.QRect(20, 50, 331, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.scannedlabel.setFont(font)
        self.scannedlabel.setAlignment(QtCore.Qt.AlignCenter)
        self.scannedlabel.setObjectName("scannedlabel")
        self.onebb = QtWidgets.QLabel(self.centralwidget)
        self.onebb.setGeometry(QtCore.QRect(20, 170, 641, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.onebb.setFont(font)
        self.onebb.setAlignment(QtCore.Qt.AlignCenter)
        self.onebb.setObjectName("onebb")
        self.twobb = QtWidgets.QLabel(self.centralwidget)
        self.twobb.setGeometry(QtCore.QRect(20, 210, 641, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.twobb.setFont(font)
        self.twobb.setAlignment(QtCore.Qt.AlignCenter)
        self.twobb.setObjectName("twobb")
        self.threebb = QtWidgets.QLabel(self.centralwidget)
        self.threebb.setGeometry(QtCore.QRect(20, 250, 641, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.threebb.setFont(font)
        self.threebb.setAlignment(QtCore.Qt.AlignCenter)
        self.threebb.setObjectName("threebb")
        self.lowbb = QtWidgets.QLabel(self.centralwidget)
        self.lowbb.setGeometry(QtCore.QRect(20, 290, 641, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lowbb.setFont(font)
        self.lowbb.setAlignment(QtCore.Qt.AlignCenter)
        self.lowbb.setObjectName("lowbb")
        self.oscannedlabel = QtWidgets.QLabel(self.centralwidget)
        self.oscannedlabel.setGeometry(QtCore.QRect(350, 50, 311, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.oscannedlabel.setFont(font)
        self.oscannedlabel.setAlignment(QtCore.Qt.AlignCenter)
        self.oscannedlabel.setObjectName("oscannedlabel")
        self.maxlabel = QtWidgets.QLabel(self.centralwidget)
        self.maxlabel.setGeometry(QtCore.QRect(20, 90, 331, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.maxlabel.setFont(font)
        self.maxlabel.setAlignment(QtCore.Qt.AlignCenter)
        self.maxlabel.setObjectName("maxlabel")
        self.actuallabel = QtWidgets.QLabel(self.centralwidget)
        self.actuallabel.setGeometry(QtCore.QRect(20, 130, 331, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.actuallabel.setFont(font)
        self.actuallabel.setAlignment(QtCore.Qt.AlignCenter)
        self.actuallabel.setObjectName("actuallabel")
        self.avgmaxlabel = QtWidgets.QLabel(self.centralwidget)
        self.avgmaxlabel.setGeometry(QtCore.QRect(350, 90, 311, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.avgmaxlabel.setFont(font)
        self.avgmaxlabel.setAlignment(QtCore.Qt.AlignCenter)
        self.avgmaxlabel.setObjectName("avgmaxlabel")
        self.avgactuallabel = QtWidgets.QLabel(self.centralwidget)
        self.avgactuallabel.setGeometry(QtCore.QRect(350, 130, 311, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.avgactuallabel.setFont(font)
        self.avgactuallabel.setAlignment(QtCore.Qt.AlignCenter)
        self.avgactuallabel.setObjectName("avgactuallabel")
        self.history = QtWidgets.QLabel(self.centralwidget)
        self.history.setGeometry(QtCore.QRect(30, 10, 631, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.history.setFont(font)
        self.history.setAlignment(QtCore.Qt.AlignCenter)
        self.history.setObjectName("history")
        self.backbutton = QtWidgets.QPushButton(self.centralwidget)
        self.backbutton.setGeometry(QtCore.QRect(600, 350, 75, 23))
        self.backbutton.setObjectName("backbutton")
        self.backbutton.clicked.connect(self.switchtomain)
        self.priceinput = QtWidgets.QLineEdit(self.centralwidget)
        self.priceinput.setGeometry(QtCore.QRect(10, 350, 41, 20))
        self.priceinput.setObjectName("priceinput")
        self.setbutton = QtWidgets.QPushButton(self.centralwidget)
        self.setbutton.setGeometry(QtCore.QRect(60, 350, 75, 23))
        self.setbutton.setObjectName("setbutton")
        self.setbutton.clicked.connect(self.setpricelimit)
        self.resetbuylists = QtWidgets.QPushButton(self.centralwidget)
        self.resetbuylists.setGeometry(QtCore.QRect(140, 350, 75, 23))
        self.resetbuylists.setObjectName("resetbuylists")
        self.resetbuylists.clicked.connect(self.resetallbuylists)
        self.resethistory = QtWidgets.QPushButton(self.centralwidget)
        self.resethistory.setGeometry(QtCore.QRect(220, 350, 75, 23))
        self.resethistory.setObjectName("resetbuylists")
        self.resethistory.clicked.connect(self.resetallhistory)
        BuyBackPlus.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(BuyBackPlus)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 690, 21))
        self.menubar.setObjectName("menubar")
        self.menuFile = QtWidgets.QMenu(self.menubar)
        self.menuFile.setObjectName("menuFile")
        BuyBackPlus.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(BuyBackPlus)
        self.statusbar.setObjectName("statusbar")
        BuyBackPlus.setStatusBar(self.statusbar)
        self.actionSettings = QtWidgets.QAction(BuyBackPlus)
        self.actionSettings.setObjectName("actionSettings")
        self.menuFile.addAction(self.actionSettings)
        self.menubar.addAction(self.menuFile.menuAction())
        self.actionSettings.triggered.connect(self.switchtosettings)
        self.scannedlabel.hide()
        self.onebb.hide()
        self.twobb.hide()
        self.threebb.hide()
        self.lowbb.hide()
        self.oscannedlabel.hide()
        self.maxlabel.hide()
        self.actuallabel.hide()
        self.avgmaxlabel.hide()
        self.avgactuallabel.hide()
        self.history.hide()
        self.backbutton.hide()
        self.setbutton.hide()
        self.priceinput.hide()
        self.resetbuylists.hide()
        self.resethistory.hide()

        self.pricelimitfile = open("pricelimit.txt", "r+")
        self.pricelimit = float(self.pricelimitfile.readline())
        self.pricelimitfile.close()

        self.tscancur = 0
        self.tscanall = self.databasesheet["B1"].value
        self.olscancur = 0
        self.olscanall = self.databasesheet["D1"].value
        self.maxbbcur = 0.00
        self.maxbball = float(self.databasesheet["B2"].value)
        self.abbcur = 0.00
        self.abball = float(self.databasesheet["B3"].value)

        self.powcur = 0.00
        self.powall = float(self.databasesheet["C5"].value)
        self.moncur = 0.00
        self.monall = float(self.databasesheet["C6"].value)
        self.zifcur = 0.00
        self.zifall = float(self.databasesheet["C7"].value)
        self.runcur = 0.00
        self.runall = float(self.databasesheet["C8"].value)
        self.camcur = 0.00
        self.camall = float(self.databasesheet["C9"].value)
        self.tbrcur = 0.00
        self.tbrall = float(self.databasesheet["C10"].value)
        self.sbyball = 0.00
        self.sbybcur = float(self.databasesheet["C11"].value)

        self.onetopbb = float(self.databasesheet["F5"].value)
        self.onetopbbname = self.databasesheet["E5"].value
        self.twotopbb = float(self.databasesheet["F6"].value)
        self.twotopbbname = self.databasesheet["E6"].value
        self.threetopbb = float(self.databasesheet["F7"].value)
        self.threetopbbname = self.databasesheet["E7"].value
        self.lowestbb = float(self.databasesheet["F8"].value)
        self.lowestbbname = self.databasesheet["E8"].value

        self.lastscanned = ""
        self.buylistscanned = 0

        self.retranslateUi(BuyBackPlus)
        QtCore.QMetaObject.connectSlotsByName(BuyBackPlus)

    def switchtosettings(self):
        self.scannedlabel.show()
        self.onebb.show()
        self.twobb.show()
        self.threebb.show()
        self.lowbb.show()
        self.oscannedlabel.show()
        self.maxlabel.show()
        self.actuallabel.show()
        self.avgmaxlabel.show()
        self.avgactuallabel.show()
        self.history.show()
        self.backbutton.show()
        self.setbutton.show()
        self.priceinput.show()
        self.resetbuylists.show()
        self.resethistory.show()

        self.updatehistorytext()

        self.titlelabel.hide()
        self.isbnlabel.hide()
        self.topbuyback.hide()
        self.firstlabel.hide()
        self.secondbuyback.hide()
        self.secondlabel.hide()
        self.thirdbuyback.hide()
        self.thirdlabel.hide()
        self.otherlabel.hide()
        self.otherbuyback.hide()
        self.offerlabel.hide()
        self.statuslabel.hide()

    def switchtomain(self):
        self.scannedlabel.hide()
        self.onebb.hide()
        self.twobb.hide()
        self.threebb.hide()
        self.lowbb.hide()
        self.oscannedlabel.hide()
        self.maxlabel.hide()
        self.actuallabel.hide()
        self.avgmaxlabel.hide()
        self.avgactuallabel.hide()
        self.history.hide()
        self.backbutton.hide()
        self.setbutton.hide()
        self.priceinput.hide()
        self.resetbuylists.hide()
        self.resethistory.hide()

        self.titlelabel.show()
        self.isbnlabel.show()
        self.topbuyback.show()
        self.firstlabel.show()
        self.secondbuyback.show()
        self.secondlabel.show()
        self.thirdbuyback.show()
        self.thirdlabel.show()
        self.otherlabel.show()
        self.otherbuyback.show()
        self.offerlabel.show()
        self.statuslabel.show()

    def setpricelimit(self):
        if float(self.priceinput.text()) > 0.00:
            self.pricelimit = float(self.priceinput.text())
            pricelimitfile = open("pricelimit.txt", "w+")
            pricelimitfile.write(str(self.pricelimit))
            pricelimitfile.close()
            self.history.setText("History: Current Session (Total) Current Price Limit: $" + str(self.pricelimit))

    def resetallbuylists(self):
        for i in range(1, 10000, 1):
            counter = 0
            if self.powelldbs["A" + str(i)].value:
                self.powelldbs["A" + str(i)].value = ""
                counter += 1
            if self.bookmonsterdbs["A" + str(i)].value:
                self.bookmonsterdbs["A" + str(i)].value = ""
                counter += 1
            if self.booksrundbs["A" + str(i)].value:
                self.booksrundbs["A" + str(i)].value = ""
                counter += 1
            if self.ecampusdbs["A" + str(i)].value:
                self.ecampusdbs["A" + str(i)].value = ""
                counter += 1
            if self.sbybdbs["A" + str(i)].value:
                self.sbybdbs["A" + str(i)].value = ""
                counter += 1
            if self.tbrecdbs["A" + str(i)].value:
                self.tbrecdbs["A" + str(i)].value = ""
                counter += 1
            if self.ziffitdbs["A" + str(i)].value:
                self.ziffitdbs["A" + str(i)].value = ""
                counter += 1
            if counter == 0:
                break

        self.powelldb.save("./buylists/powell.xlsx")
        self.bookmonsterdb.save("./buylists/bookmonster.xlsx")
        self.booksrundb.save("./buylists/booksrun.xlsx")
        self.ecampusdb.save("./buylists/ecampus.xlsx")
        self.sbybdb.save("./buylists/sbyb.xlsx")
        self.tbrecdb.save("./buylists/tbrec.xlsx")
        self.ziffitdb.save("./buylists/ziffit.xlsx")
        print("Buylists are Emptied!")

    def resetallhistory(self):
        self.tscancur = 0
        self.tscanall = 0
        self.olscancur = 0
        self.olscanall = 0
        self.maxbbcur = 0.00
        self.maxbball = 0.00
        self.abbcur = 0.00
        self.abball = 0.00
        self.powcur = 0.00
        self.powall = 0.00
        self.moncur = 0.00
        self.monall = 0.00
        self.zifcur = 0.00
        self.zifall = 0.00
        self.runcur = 0.00
        self.runall = 0.00
        self.camcur = 0.00
        self.camall = 0.00
        self.tbrcur = 0.00
        self.tbrall = 0.00
        self.sbyball = 0.00
        self.sbybcur = 0.00
        self.onetopbb = 0.00
        self.onetopbbname = ""
        self.twotopbb = 0.00
        self.twotopbbname = ""
        self.threetopbb = 0.00
        self.threetopbbname = ""
        self.lowestbb = 0.00
        self.lowestbbname = ""
        updatehistory()
        self.updatehistorytext()
        print("History has been reset!")

    def updatehistorytext(self):
        self.scannedlabel.setText("Total Scanned: " + str(self.tscancur) + " (" + str(self.tscanall) + ")")
        self.maxlabel.setText("Est. Max BB Total: $" + str(self.maxbbcur) + " ($" + str(self.maxbball) + ")")
        self.actuallabel.setText("Est. Actual BB Total: $" + str(self.abbcur) + " ($" + str(self.abball) + ")")
        self.oscannedlabel.setText("Over Limit Scanned: " + str(self.olscancur) + " (" + str(self.olscanall) + ")")
        if (self.olscancur > 0 and self.olscanall > 0):
            self.avgmaxlabel.setText(
                "Est. Avg Max BB Per Item: $" + str(float(self.maxbbcur / self.olscancur))[:4] + " ($" + str(
                    float(self.maxbball / self.olscanall))[:4] + ")")
            self.avgactuallabel.setText(
                "Est. Avg Actual BB Per Item: $" + str(float(self.abbcur / self.olscancur))[:4] + " ($" + str(
                    float(self.abball / self.olscanall))[:4] + ")")
        else:
            self.avgmaxlabel.setText("Est. Avg Max BB Per Item: $0.0 ($0.0)")
            self.avgactuallabel.setText("Est. Avg Actual BB Per Item: $0.0 ($0.0)")
        self.onebb.setText("Top BB Site Overall: " + self.onetopbbname + " $" + str(self.onetopbb))
        self.twobb.setText("Second Top BB Site Overall: " + self.twotopbbname + " $" + str(self.twotopbb))
        self.threebb.setText("Third Top BB Site Overall: " + self.threetopbbname + " $" + str(self.threetopbb))
        self.lowbb.setText("Lowest BB Site Overall: " + self.lowestbbname + " $" + str(self.lowestbb))

    def retranslateUi(self, BuyBackPlus):
        _translate = QtCore.QCoreApplication.translate
        BuyBackPlus.setWindowTitle(_translate("BuyBackPlus", "BuyBackPlus"))
        self.titlelabel.setText(_translate("BuyBackPlus", "Title: None"))
        self.isbnlabel.setText(_translate("BuyBackPlus", "ISBN Scanned: None"))
        self.topbuyback.setText(_translate("BuyBackPlus", "TOP BUYBACK SITE"))
        self.firstlabel.setText(_translate("BuyBackPlus", "None"))
        self.secondbuyback.setText(_translate("BuyBackPlus", "2ND TOP BUYBACK SITE"))
        self.secondlabel.setText(_translate("BuyBackPlus", "None"))
        self.thirdbuyback.setText(_translate("BuyBackPlus", "3RD TOP BUYBACK SITE"))
        self.thirdlabel.setText(_translate("BuyBackPlus", "None"))
        self.otherlabel.setText(_translate("BuyBackPlus", "None"))
        self.otherbuyback.setText(_translate("BuyBackPlus", "OTHER BUYBACK SITES"))
        self.offerlabel.setText(_translate("BuyBackPlus", "Are Lower Offers Available : N/A"))
        self.statuslabel.setText(_translate("BuyBackPlus", "Status: Program Sucessfully Loaded Please Scan A Barcode!"))

        self.scannedlabel.setText(_translate("BuyBackPlus", "Total Scanned: N/A"))
        self.onebb.setText(_translate("BuyBackPlus", "Top BB Site: N/A"))
        self.twobb.setText(_translate("BuyBackPlus", "2nd Top BB Site: N/A"))
        self.threebb.setText(_translate("BuyBackPlus", "3rd Top BB Site: N/A"))
        self.lowbb.setText(_translate("BuyBackPlus", "Lowest BB Site: N/A"))
        self.oscannedlabel.setText(_translate("BuyBackPlus", "Over Limit Scanned: N/A"))
        self.maxlabel.setText(_translate("BuyBackPlus", "Est. Max BB Total: N/A"))
        self.actuallabel.setText(_translate("BuyBackPlus", "Est. Actual BB Total: N/A"))
        self.avgmaxlabel.setText(_translate("BuyBackPlus", "Est. Avg BB Per Item: N/A"))
        self.avgactuallabel.setText(_translate("BuyBackPlus", "Est. Avg Actual BB Per Item: N/A"))
        self.history.setText(_translate("BuyBackPlus", "History: Current Session (Total) Current Price Limit: $" + str(self.pricelimit)))
        self.backbutton.setText(_translate("BuyBackPlus", "Back"))
        self.setbutton.setText(_translate("BuyBackPlus", "Set Price Limit"))
        self.resetbuylists.setText(_translate("BuyBackPlus", "Reset Buylists"))
        self.resethistory.setText(_translate("BuyBackPlus", "Reset History"))
        self.menuFile.setTitle(_translate("BuyBackPlus", "File"))
        self.actionSettings.setText(_translate("BuyBackPlus", "Settings"))
        self.actionSettings.setStatusTip(_translate("BuyBackPlus", "Set price limit and view stats"))
        self.actionSettings.setShortcut(_translate("BuyBackPlus", "Ctrl+S"))

isbn = ""
isbnlist = ""
count = 0
totaltime = 0.00
lastbbone = 0.00
lastbbonename = ""
lastbbtwo = 0.00
lastbbtwoname = ""
lastbbthree = 0.00
lastbbthreename = ""
maxbb = 0.00

bbadd = 0.00
bbaddname = ""
savetitle = ""

session = requests.Session()

jar = requests.cookies.RequestsCookieJar()
jar.set("ga", "GA1.2.743815479.1563760638; __stripe_mid=5b6716ae-9fa6-4e76-a5d7-665d1deb4430")
jar.set("fbp", "fb.1.1573365037498.637495270; ref45a8bc6=79ac2213350174f394a53f2a41cb57e0")

session.cookies = jar
hdr = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.79 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    'Accept-Encoding': 'gzip, deflate, br',
    'Accept-Language': 'en-US,en;q=0.9',
    'Connection': 'keep-alive'}

def updatetopbb():
    one = ui.powall
    onename = "Powell's"
    two = ui.monall
    twoname = "BookMonster"
    three = ui.zifall
    threename = "Ziffit"
    four = ui.sbyball
    fourname = "SellBackYourBook"
    five = ui.tbrall
    fivename = "TextBookRecycling"
    six = ui.camall
    sixname = "eCampus"
    seven = ui.runall
    sevenname = "BooksRun"

    spotone = one
    spotonename = onename
    spottwo = 0.00
    spottowname = "N/A"
    spotthree = 0.00
    spotthreename = "N/A"
    spotlast = 0.00
    spotlastname = "N/A"

    if two > spotone:
        spottwo = spotone
        spottwoname = spotonename
        spotone = two
        spotonename = twoname
    else:
        spottwo = two
        spottwoname = twoname

    if three > spotone:
        spotthree = spottwo
        spotthreename = spottowname
        spottwo = spotone
        spottwoname = spotonename
        spotone = three
        spotonename = threename
    elif three > spottwo:
        spotthree = spottwo
        spotthree = spottwoname
        spottwo = three
        spottwo = threename
    else:
        spotthree = three
        spotthreename = threename

    if four > spotone:
        spotlast = spotthree
        spotlastname = spotthreename
        spotthree = spottwo
        spotthreename = spottowname
        spottwo = spotone
        spottwoname = spotonename
        spotone = four
        spotonename = fourname
    elif four > spottwo:
        spotlast = spotthree
        spotlastname = spotthreename
        spotthree = spottwo
        spotthree = spottwoname
        spottwo = four
        spottwo = fourname
    elif four > spotthree:
        spotlast = spotthree
        spotlastname = spotthreename
        spotthree = four
        spotthreename = fourname
    else:
        spotlast = four
        spotlastname = fourname

    if five > spotone:
        spotthree = spottwo
        spotthreename = spottowname
        spottwo = spotone
        spottwoname = spotonename
        spotone = five
        spotonename = fivename
    elif five > spottwo:
        spotthree = spottwo
        spotthree = spottwoname
        spottwo = five
        spottwo = fivename
    elif five > spotthree:
        spotlast = spotthree
        spotlastname = spotthreename
        spotthree = five
        spotthreename = fivename
    elif five < spotlast:
        spotlast = five
        spotlastname = fivename

    if six > spotone:
        spotthree = spottwo
        spotthreename = spottowname
        spottwo = spotone
        spottwoname = spotonename
        spotone = six
        spotonename = sixname
    elif six > spottwo:
        spotthree = spottwo
        spotthree = spottwoname
        spottwo = six
        spottwo = sixname
    elif six > spotthree:
        spotlast = spotthree
        spotlastname = spotthreename
        spotthree = six
        spotthreename = sixname
    elif six < spotlast:
        spotlast = six
        spotlastname = sixname

    if seven > spotone:
        spotthree = spottwo
        spotthreename = spottowname
        spottwo = spotone
        spottwoname = spotonename
        spotone = seven
        spotonename = sevenname
    elif seven > spottwo:
        spotthree = spottwo
        spotthree = spottwoname
        spottwo = seven
        spottwo = sevenname
    elif seven > spotthree:
        spotlast = spotthree
        spotlastname = spotthreename
        spotthree = seven
        spotthreename = sevenname
    elif seven < spotlast:
        spotlast = five
        spotlastname = sevenname

    ui.onetopbb = spotone
    ui.onetopbbname = spotonename
    ui.twotopbb = spottwo
    ui.twotopbbname = spottwoname
    ui.threetopbb = spotthree
    ui.threetopbbname = spotthreename
    ui.lowestbb = spotlast
    ui.lowestbbname = spotlastname

def updatehistory():
    updatetopbb()
    ui.databasesheet["B1"] = ui.tscanall
    ui.databasesheet["B2"] = ui.maxbball
    ui.databasesheet["B3"] = ui.abball
    ui.databasesheet["D1"] = ui.olscanall
    if(ui.olscanall > 0):
        ui.databasesheet["D2"] = str(float(ui.maxbball / ui.olscanall))[:4]
        ui.databasesheet["D3"] = str(float(ui.abball / ui.olscanall))[:4]
    else:
        ui.databasesheet["D2"] = 0
        ui.databasesheet["D3"] = 0

    ui.databasesheet["C5"] = ui.powall
    ui.databasesheet["C6"] = ui.monall
    ui.databasesheet["C7"] = ui.zifall
    ui.databasesheet["C8"] = ui.runall
    ui.databasesheet["C9"] = ui.camall
    ui.databasesheet["C10"] = ui.tbrall
    ui.databasesheet["C11"] = ui.sbyball

    ui.databasesheet["E5"] = ui.onetopbbname
    ui.databasesheet["F5"] = ui.onetopbb
    ui.databasesheet["E6"] = ui.twotopbbname
    ui.databasesheet["F6"] = ui.twotopbb
    ui.databasesheet["E7"] = ui.threetopbbname
    ui.databasesheet["F7"] = ui.threetopbb
    ui.databasesheet["E8"] = ui.lowestbbname
    ui.databasesheet["F8"] = ui.lowestbb

    ui.database.save("./history.xlsx")
    excelread = pd.read_excel("./history.xlsx")
    excelread.to_csv("./history.tsv", sep="\t", index=False)

def exit_handler():
    updatehistory()
    ui.database.save("./history.xlsx")
    ui.powelldb.save("./buylists/powell.xlsx")
    ui.bookmonsterdb.save("./buylists/bookmonster.xlsx")
    ui.booksrundb.save("./buylists/booksrun.xlsx")
    ui.ecampusdb.save("./buylists/ecampus.xlsx")
    ui.sbybdb.save("./buylists/sbyb.xlsx")
    ui.tbrecdb.save("./buylists/tbrec.xlsx")
    ui.ziffitdb.save("./buylists/ziffit.xlsx")

atexit.register(exit_handler)

def on_press(key):
    global isbn
    global isbnlist
    global isbnarray
    global count
    global totaltime
    global lastbbone
    global lastbbonename
    global maxbb
    global lastbbtwo
    global lastbbtwoname
    global lastbbthree
    global lastbbthreenam
    global bbadd
    global bbaddname
    global savetitle
    try:
        if key == Key.enter:
            if (len(isbn) == 13) or (len(isbn) == 10):
                bblist = False
                if isbn == "6548512751238":
                    bbadd = lastbbone
                    bbaddname = lastbbonename
                    bblist = True
                elif isbn == "6548485651238":
                    bbadd = lastbbtwo
                    bbaddname = lastbbtwoname
                    bblist = True
                elif isbn == "6548662341238":
                    bbadd = lastbbthree
                    bbaddname = lastbbthreename
                    bblist = True
                if bblist:
                    error = False
                    placeholdercur = float(ui.abbcur + bbadd)
                    placeholderall = float(ui.abball + bbadd)
                    ui.abbcur = placeholdercur
                    ui.abball = placeholderall
                    if bbaddname == "Powell's":
                        for i in range(1000):
                            cellnum = str(i + 1)
                            if not ui.powelldbs["A" + cellnum].value:
                                #ui.powelldbs["A" + cellnum] = savetitle
                                #ui.powelldbs["B" + cellnum] = ui.lastscanned
                                #ui.powelldbs["C" + cellnum] = "$" + str(bbadd)
                                ui.powelldbs["A" + cellnum] = ui.lastscanned
                                ui.powcur += bbadd
                                ui.powall += bbadd
                                ui.powelldb.save("./buylists/powell.xlsx")
                                excelread = pd.read_excel("./buylists/powell.xlsx")
                                excelread.to_csv("./buylists/powell.tsv", sep="\t", index=False)
                                break
                    elif bbaddname == "BookMonster":
                        for i in range(1000):
                            cellnum = str(i + 1)
                            if not ui.bookmonsterdbs["A" + cellnum].value:
                                #ui.bookmonsterdbs["A" + cellnum] = savetitle
                                #ui.bookmonsterdbs["B" + cellnum] = ui.lastscanned
                                #ui.bookmonsterdbs["C" + cellnum] = "$" + str(bbadd)
                                ui.bookmonsterdbs["A" + cellnum] = ui.lastscanned
                                ui.moncur += bbadd
                                ui.monall += bbadd
                                ui.bookmonsterdb.save("./buylists/bookmonster.xlsx")
                                excelread = pd.read_excel("./buylists/bookmonster.xlsx")
                                excelread.to_csv("./buylists/bookmonster.tsv", sep="\t", index=False)
                                break
                    elif bbaddname == "Ziffit":
                        for i in range(1000):
                            cellnum = str(i + 1)
                            if not ui.ziffitdbs["A" + cellnum].value:
                                #ui.ziffitdbs["A" + cellnum] = savetitle
                                #ui.ziffitdbs["B" + cellnum] = ui.lastscanned
                                #ui.ziffitdbs["C" + cellnum] = "$" + str(bbadd)
                                ui.ziffitdbs["A" + cellnum] = ui.lastscanned
                                ui.zifcur += bbadd
                                ui.zifall += bbadd
                                ui.ziffitdb.save("./buylists/ziffit.xlsx")
                                excelread = pd.read_excel("./buylists/ziffit.xlsx")
                                excelread.to_csv("./buylists/ziffit.tsv", sep="\t", index=False)
                                break
                    elif bbaddname == "BooksRun":
                        for i in range(1000):
                            cellnum = str(i + 1)
                            if not ui.booksrundbs["A" + cellnum].value:
                                #ui.booksrundbs["A" + cellnum] = savetitle
                                #ui.booksrundbs["B" + cellnum] = ui.lastscanned
                                #ui.booksrundbs["C" + cellnum] = "$" + str(bbadd)
                                ui.booksrundbs["A" + cellnum] = ui.lastscanned
                                ui.runcur += bbadd
                                ui.runall += bbadd
                                ui.booksrundb.save("./buylists/booksrun.xlsx")
                                excelread = pd.read_excel("./buylists/booksrun.xlsx")
                                excelread.to_csv("./buylists/booksrun.tsv", sep="\t", index=False)
                                break
                    elif bbaddname == "eCampus":
                        for i in range(1000):
                            cellnum = str(i + 1)
                            if not ui.ecampusdbs["A" + cellnum].value:
                                #ui.ecampusdbs["A" + cellnum] = savetitle
                                #ui.ecampusdbs["B" + cellnum] = ui.lastscanned
                                #ui.ecampusdbs["C" + cellnum] = "$" + str(bbadd)
                                ui.ecampusdbs["A" + cellnum] = ui.lastscanned
                                ui.camcur += bbadd
                                ui.camall += bbadd
                                ui.ecampusdb.save("./buylists/ecampus.xlsx")
                                excelread = pd.read_excel("./buylists/ecampus.xlsx")
                                excelread.to_csv("./buylists/ecampus.tsv", sep="\t", index=False)
                                break
                    elif bbaddname == "TextbookRecycling":
                        for i in range(1000):
                            cellnum = str(i + 1)
                            if not ui.tbrecdbs["A" + cellnum].value:
                                #ui.tbrecdbs["A" + cellnum] = savetitle
                                #ui.tbrecdbs["B" + cellnum] = ui.lastscanned
                                #ui.tbrecdbs["C" + cellnum] = "$" + str(bbadd)
                                ui.tbrecdbs["A" + cellnum] = ui.lastscanned
                                ui.tbrcur += bbadd
                                ui.tbral += bbadd
                                ui.tbrecdb.save("./buylists/tbrec.xlsx")
                                excelread = pd.read_excel("./buylists/tbrec.xlsx")
                                excelread.to_csv("./buylists/tbrec.tsv", sep="\t", index=False)
                                break
                    elif bbaddname == "SellBackYourBook":
                        for i in range(1000):
                            cellnum = str(i + 1)
                            if not ui.sbybdbs["A" + cellnum].value:
                                #ui.sbybdbs["A" + cellnum] = savetitle
                                #ui.sbybdbs["B" + cellnum] = ui.lastscanned
                                #ui.sbybdbs["C" + cellnum] = "$" + str(bbadd)
                                ui.sbybdbs["A" + cellnum] = ui.lastscanned
                                ui.sbybcur += bbadd
                                ui.sbyball += bbadd
                                ui.sbybdb.save("./buylists/sbyb.xlsx")
                                excelread = pd.read_excel("./buylists/sbyb.xlsx")
                                excelread.to_csv("./buylists/sbyb.tsv", sep="\t", index=False)
                                break
                    else:
                        error = True
                        ui.statuslabel.setText("Status: ERROR This buyback is currently not supported sorry!")
                    if not error:
                        ui.statuslabel.setText("Status: ISBN " + ui.lastscanned + " has been added to the " + bbaddname + " buy list!")
                        isbn = ""
                else:
                    print(isbn)
                    print("Bookscouter Enter")
                    start = time.time()
                    findbookprices(isbn)
                    end = time.time()
                    print(str(end - start) + " Seconds")
                    print("Bookscouter Exit")
                    count += 1
                    totaltime += end - start
                    ui.statuslabel.setText("Status: " + "Load Time: " + str(end - start)[:4] + " Avg Load Time: " + str(totaltime / count)[:4] + " Seconds")
                    ui.maxbbcur += maxbb
                    ui.maxbball += maxbb
                    maxbb = 0.00
                updatehistory()
                ui.centralwidget.update()
            else:
                ui.titlelabel.setText("Invalid")
                ui.firstlabel.setText("Invalid")
                ui.secondlabel.setText("Invalid")
                ui.thirdlabel.setText("Invalid")
                ui.otherlabel.setText("Invalid")
                winsound.PlaySound("no.wav", winsound.SND_ASYNC)
            if isbn != "6548512751238" and isbn != "6548485651238" and isbn != "6548662341238":
                print(isbn)
                ui.isbnlabel.setText("ISBN Scanned: " + isbn)
                ui.lastscanned = isbn
            else:
                ui.lastscanned = ""
                isbn = ""
            isbn = ""
            return
        placeholder = str(key)
        placeholdertwo = placeholder.replace("'", "")
        placeholderthree = placeholdertwo.replace("<", "")
        check = int(placeholderthree.replace(">", ""))
        if check >= 96:
            if check <= 105:
                check -= 96
        if check >= 0:
            if check <= 9:
                isbn += str(check)
    except AttributeError:
        return key
    except TypeError:
        return key
    except ValueError:
        return key

def findbookprices(isbn):
    global lastbbone
    global lastbbonename
    global lastbbtwo
    global lastbbtwoname
    global lastbbthree
    global lastbbthreename
    global savetitle
    global maxbb
    global session
    global hdr

    url = "https://api.bookscouter.com/v3/prices/sell/" + isbn + "?vendors=all"
    session.get(url, headers=hdr)
    response = session.get(url, headers=hdr)
    info = json.loads(response.text)
    print(info)

    try:
        data = info['data']
        pricelist = data['Prices']
        book = data['Book']
        title = str(book['Title'])
        savetitle = title
        buying = False
        loweroffer = False
        counter = 1
        firstbb = "Not Buying/Price Too Low"
        secondbb = "Not Buying/Price Too Low"
        thirdbb = "Not Buying/Price Too Low"
        listother = "Not Buying/Price Too Low"
        for Price in pricelist:
            price = Price['Price']
            floprice = float(price)
            if floprice > 0:
                if floprice > 0 and floprice < ui.pricelimit:
                    loweroffer = True
                if floprice > ui.pricelimit:
                    buying = True
                    vendor = Price['Vendor']
                    name = vendor['Name']
                    if counter == 1:
                        firstbb = name + " $" + str(price)
                        lastbbone = floprice
                        lastbbonename = name
                        maxbb = floprice
                        counter += 1
                    elif counter == 2:
                        secondbb = name + " $" + str(price)
                        lastbbtwo = floprice
                        lastbbtwoname = name
                        counter += 1
                    elif counter == 3:
                        thirdbb = name + " $" + str(price)
                        lastbbthree = floprice
                        lastbbthreename = name
                        counter += 1
                    elif counter == 4:
                        listother = name + " $" + str(price) + "\n"
                        counter += 1
                    else:
                        listother += name + " $" + str(price) + "\n"
            else:
                break
        if loweroffer:
            ui.offerlabel.setText("Are Lower Offers Available: Yes!")
            ui.offerlabel.setStyleSheet("background: green")
        else:
            ui.offerlabel.setText("Are Lower Offers Available: No :(")
            ui.offerlabel.setStyleSheet("background: red")
        if buying:
            ui.titlelabel.setText("Title: " + title[:23])
            ui.firstlabel.setText(firstbb)
            ui.secondlabel.setText(secondbb)
            ui.thirdlabel.setText(thirdbb)
            ui.otherlabel.setText(listother)
            winsound.PlaySound("yes.wav", winsound.SND_ASYNC)
            ui.olscancur += 1
            ui.olscanall += 1
        else:
            ui.titlelabel.setText("Title: " + title[:23])
            ui.firstlabel.setText(firstbb)
            ui.secondlabel.setText(secondbb)
            ui.thirdlabel.setText(thirdbb)
            ui.otherlabel.setText(listother)
            winsound.PlaySound("deleted.wav", winsound.SND_ASYNC)
        ui.tscancur += 1
        ui.tscanall += 1
    except KeyError:
        print(info['error'])
        ui.titlelabel.setText("Invalid")
        ui.firstlabel.setText("Invalid")
        ui.secondlabel.setText("Invalid")
        ui.thirdlabel.setText("Invalid")
        ui.otherlabel.setText("Invalid")
        winsound.PlaySound("no.wav", winsound.SND_ASYNC)
    except TypeError:
        print("TypeError")
        ui.titlelabel.setText("Invalid")
        ui.firstlabel.setText("Invalid")
        ui.secondlabel.setText("Invalid")
        ui.thirdlabel.setText("Invalid")
        ui.otherlabel.setText("Invalid")
        winsound.PlaySound("no.wav", winsound.SND_ASYNC)

def startlistener():
    with Listener(on_press=on_press) as listener:
        listener.join()

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    BuyBackPlus = QtWidgets.QMainWindow()
    ui = Ui_BuyBackPlus()
    ui.setupUi(BuyBackPlus)
    BuyBackPlus.show()
    listenthread = threading.Thread(target=startlistener, args=())
    listenthread.start()
    sys.exit(app.exec_())
